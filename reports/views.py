from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime, timedelta
from products.models import Product
from inventory.models import Import, Export, Batch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import json

def adjust_column_width(worksheet):
    """Điều chỉnh độ rộng cột cho worksheet"""
    for column in worksheet.columns:
        max_length = 0
        try:
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        except AttributeError:
            # Bỏ qua nếu là merged cell
            continue

@login_required
def report_list(request):
    """Danh sách các báo cáo"""
    return render(request, 'reports/report_list.html')

@login_required
def inventory_report(request):
    """Báo cáo tồn kho"""
    products = Product.objects.filter(is_active=True)
    
    # Thống kê tổng quan
    total_products = products.count()
    low_stock_products = []
    expiring_products = []
    
    # Dữ liệu chi tiết cho bảng
    inventory_details = []
    today = timezone.now().date()
    expiring_soon = today + timedelta(days=270)  # 9 tháng
    
    for product in products:
        # Kiểm tra sản phẩm sắp hết hàng (tổng tồn kho <= 1)
        if product.total_stock <= 1:
            low_stock_products.append(product)
        
        # Lấy tất cả lô hàng của sản phẩm
        batches = product.batches.filter(is_active=True, remaining_quantity__gt=0).order_by('import_date')
        
        for batch in batches:
            # Kiểm tra lô hàng sắp hết hạn (expiry_date là property)
            if batch.expiry_date and batch.expiry_date <= expiring_soon:
                expiring_products.append(product)
            
            # Thêm vào danh sách chi tiết
            inventory_details.append({
                'product': product,
                'batch': batch,
                'status': 'normal'
            })
    
    # Loại bỏ trùng lặp trong expiring_products
    expiring_products = list(set(expiring_products))
    
    context = {
        'total_products': total_products,
        'low_stock_products': low_stock_products,
        'expiring_products': expiring_products,
        'inventory_details': inventory_details,
        'today': today,
        'expiring_soon': expiring_soon,
    }
    return render(request, 'reports/inventory_report.html', context)

@login_required
def import_export_report(request):
    """Báo cáo nhập xuất kho"""
    # Lọc theo thời gian
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    imports = Import.objects.all()
    exports = Export.objects.all()
    
    if start_date:
        imports = imports.filter(import_date__gte=start_date)
        exports = exports.filter(export_date__gte=start_date)
    
    if end_date:
        imports = imports.filter(import_date__lte=end_date)
        exports = exports.filter(export_date__lte=end_date)
    
    # Tính tổng
    total_import_value = sum(imp.total_amount for imp in imports)
    total_export_value = sum(exp.total_amount for exp in exports)
    profit = total_export_value - total_import_value
    
    context = {
        'imports': imports,
        'exports': exports,
        'total_import_value': total_import_value,
        'total_export_value': total_export_value,
        'profit': profit,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'reports/import_export_report.html', context)

@login_required
def export_inventory_excel(request):
    """Xuất báo cáo tồn kho ra Excel"""
    # Lấy dữ liệu tương tự như inventory_report
    products = Product.objects.filter(is_active=True)
    
    # Tạo workbook mới
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo tồn kho"
    
    # Định dạng header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Tiêu đề báo cáo
    ws.merge_cells('A1:J1')
    ws['A1'] = "BÁO CÁO TỒN KHO CHI TIẾT"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Header bảng
    headers = ['STT', 'Mã SP', 'Tên sản phẩm', 'Danh mục', 'Đơn vị', 'Tổng tồn kho', 'Lô hàng', 'Hạn sử dụng', 'Số lượng còn', 'Trạng thái']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Dữ liệu
    row = 4
    today = timezone.now().date()
    expiring_soon = today + timedelta(days=270)  # 9 tháng
    
    for product in products:
        batches = product.batches.filter(is_active=True, remaining_quantity__gt=0).order_by('import_date')
        
        for batch in batches:
            # Xác định trạng thái (expiry_date là property)
            if batch.expiry_date and batch.expiry_date < today:
                status = "Hết hạn"
            elif batch.expiry_date and batch.expiry_date < expiring_soon:
                status = "Sắp hết hạn"
            elif batch.remaining_quantity <= 1:
                status = "Sắp hết hàng"
            else:
                status = "Bình thường"
            
            # Ghi dữ liệu
            ws.cell(row=row, column=1, value=row-3)
            ws.cell(row=row, column=2, value=product.code)
            ws.cell(row=row, column=3, value=product.name)
            ws.cell(row=row, column=4, value=product.category.name)
            ws.cell(row=row, column=5, value=product.unit)
            ws.cell(row=row, column=6, value=product.total_stock)
            ws.cell(row=row, column=7, value=batch.batch_code)
            ws.cell(row=row, column=8, value=batch.expiry_date.strftime('%d/%m/%Y') if batch.expiry_date else '-')
            ws.cell(row=row, column=9, value=batch.remaining_quantity)
            ws.cell(row=row, column=10, value=status)
            row += 1
    
    # Điều chỉnh độ rộng cột
    adjust_column_width(ws)
    
    # Tạo response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="bao_cao_ton_kho_{today.strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def export_import_export_excel(request):
    """Xuất báo cáo nhập/xuất kho ra Excel"""
    # Lấy tham số lọc
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Lọc dữ liệu
    imports = Import.objects.all()
    exports = Export.objects.all()
    
    if start_date:
        imports = imports.filter(import_date__gte=start_date)
        exports = exports.filter(export_date__gte=start_date)
    
    if end_date:
        imports = imports.filter(import_date__lte=end_date)
        exports = exports.filter(export_date__lte=end_date)
    
    # Tạo workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Phiếu nhập
    ws1 = wb.active
    ws1.title = "Phiếu nhập kho"
    
    # Tiêu đề
    ws1.merge_cells('A1:G1')
    ws1['A1'] = "BÁO CÁO PHIẾU NHẬP KHO"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A1'].alignment = Alignment(horizontal="center")
    
    # Header
    headers = ['STT', 'Mã phiếu', 'Ngày nhập', 'Nhà cung cấp', 'Người tạo', 'Tổng giá trị', 'Ghi chú']
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Dữ liệu phiếu nhập
    row = 4
    for import_order in imports:
        ws1.cell(row=row, column=1, value=row-3)
        ws1.cell(row=row, column=2, value=import_order.import_code)
        ws1.cell(row=row, column=3, value=import_order.import_date.strftime('%d/%m/%Y %H:%M'))
        ws1.cell(row=row, column=4, value=import_order.supplier or '-')
        ws1.cell(row=row, column=5, value=import_order.created_by.get_full_name() or import_order.created_by.username)
        ws1.cell(row=row, column=6, value=import_order.total_amount)
        ws1.cell(row=row, column=7, value=import_order.notes or '-')
        row += 1
    
    # Sheet 2: Phiếu xuất
    ws2 = wb.create_sheet("Phiếu xuất kho")
    
    # Tiêu đề
    ws2.merge_cells('A1:G1')
    ws2['A1'] = "BÁO CÁO PHIẾU XUẤT KHO"
    ws2['A1'].font = Font(bold=True, size=16)
    ws2['A1'].alignment = Alignment(horizontal="center")
    
    # Header
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Dữ liệu phiếu xuất
    row = 4
    for export_order in exports:
        ws2.cell(row=row, column=1, value=row-3)
        ws2.cell(row=row, column=2, value=export_order.export_code)
        ws2.cell(row=row, column=3, value=export_order.export_date.strftime('%d/%m/%Y %H:%M'))
        ws2.cell(row=row, column=4, value=export_order.customer or '-')
        ws2.cell(row=row, column=5, value=export_order.created_by.get_full_name() or export_order.created_by.username)
        ws2.cell(row=row, column=6, value=export_order.total_amount)
        ws2.cell(row=row, column=7, value=export_order.notes or '-')
        row += 1
    
    # Sheet 3: Tổng hợp
    ws3 = wb.create_sheet("Tổng hợp")
    
    # Tiêu đề
    ws3.merge_cells('A1:D1')
    ws3['A1'] = "TỔNG HỢP NHẬP/XUẤT KHO"
    ws3['A1'].font = Font(bold=True, size=16)
    ws3['A1'].alignment = Alignment(horizontal="center")
    
    # Thống kê
    total_import_value = sum(imp.total_amount for imp in imports)
    total_export_value = sum(exp.total_amount for exp in exports)
    profit = total_export_value - total_import_value
    
    stats = [
        ['Chỉ tiêu', 'Giá trị'],
        ['Tổng phiếu nhập', len(imports)],
        ['Tổng phiếu xuất', len(exports)],
        ['Tổng giá trị nhập', total_import_value],
        ['Tổng giá trị xuất', total_export_value],
        ['Lợi nhuận', profit],
    ]
    
    for row_idx, (label, value) in enumerate(stats, 3):
        ws3.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws3.cell(row=row_idx, column=2, value=value)
    
    # Điều chỉnh độ rộng cột cho tất cả sheets
    for ws in [ws1, ws2, ws3]:
        adjust_column_width(ws)
    
    # Tạo response
    today = timezone.now().date()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="bao_cao_nhap_xuat_{today.strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response

@login_required
def profit_report(request):
    """Báo cáo lợi nhuận"""
    # Lọc theo thời gian
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Lấy dữ liệu xuất kho
    exports = Export.objects.all()
    if start_date:
        exports = exports.filter(export_date__gte=start_date)
    if end_date:
        exports = exports.filter(export_date__lte=end_date)
    
    # Tính toán lợi nhuận theo sản phẩm
    profit_details = []
    total_import_value = 0
    total_export_value = 0
    
    # Lấy tất cả sản phẩm có xuất kho
    products_with_exports = set()
    for export in exports:
        for item in export.items.all():
            products_with_exports.add(item.batch.product)
    
    for product in products_with_exports:
        # Tính tổng xuất của sản phẩm
        product_exports = []
        for export in exports:
            for item in export.items.all():
                if item.batch.product == product:
                    product_exports.append(item)
        
        if product_exports:
            total_quantity = sum(item.quantity for item in product_exports)
            total_export_value_product = sum(item.total_price for item in product_exports)
            avg_export_price = total_export_value_product / total_quantity if total_quantity > 0 else 0
            
            # Tính giá nhập trung bình từ các lô hàng đã xuất
            import_prices = []
            for item in product_exports:
                import_prices.append(item.batch.import_price)
            
            avg_import_price = sum(import_prices) / len(import_prices) if import_prices else 0
            
            # Tính lợi nhuận
            profit_per_unit = avg_export_price - avg_import_price
            total_profit_product = profit_per_unit * total_quantity
            profit_margin = (profit_per_unit / avg_import_price * 100) if avg_import_price > 0 else 0
            
            profit_details.append({
                'product': product,
                'total_quantity': total_quantity,
                'avg_import_price': avg_import_price,
                'avg_export_price': avg_export_price,
                'profit_per_unit': profit_per_unit,
                'total_profit': total_profit_product,
                'profit_margin': profit_margin
            })
            
            total_export_value += total_export_value_product
            total_import_value += avg_import_price * total_quantity
    
    # Sắp xếp theo lợi nhuận giảm dần
    profit_details.sort(key=lambda x: x['total_profit'], reverse=True)
    
    # Tính tổng lợi nhuận
    total_profit = total_export_value - total_import_value
    profit_margin_total = (total_profit / total_import_value * 100) if total_import_value > 0 else 0
    
    # Dữ liệu cho biểu đồ
    top_products_labels = [item['product'].name for item in profit_details[:5]]
    top_products_profits = [float(item['total_profit']) for item in profit_details[:5]]
    
    context = {
        'profit_details': profit_details,
        'total_import_value': float(total_import_value),
        'total_export_value': float(total_export_value),
        'total_profit': float(total_profit),
        'profit_margin': float(profit_margin_total),
        'top_products_labels': json.dumps(top_products_labels, ensure_ascii=False),
        'top_products_profits': json.dumps(top_products_profits),
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'reports/profit_report.html', context)

@login_required
def export_profit_excel(request):
    """Xuất báo cáo lợi nhuận ra Excel"""
    # Lấy tham số lọc
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Lọc dữ liệu
    exports = Export.objects.all()
    if start_date:
        exports = exports.filter(export_date__gte=start_date)
    if end_date:
        exports = exports.filter(export_date__lte=end_date)
    
    # Tính toán lợi nhuận (tương tự như profit_report)
    profit_details = []
    total_import_value = 0
    total_export_value = 0
    
    products_with_exports = set()
    for export in exports:
        for item in export.items.all():
            products_with_exports.add(item.batch.product)
    
    for product in products_with_exports:
        product_exports = []
        for export in exports:
            for item in export.items.all():
                if item.batch.product == product:
                    product_exports.append(item)
        
        if product_exports:
            total_quantity = sum(item.quantity for item in product_exports)
            total_export_value_product = sum(item.total_price for item in product_exports)
            avg_export_price = total_export_value_product / total_quantity if total_quantity > 0 else 0
            
            import_prices = []
            for item in product_exports:
                import_prices.append(item.batch.import_price)
            
            avg_import_price = sum(import_prices) / len(import_prices) if import_prices else 0
            
            profit_per_unit = avg_export_price - avg_import_price
            total_profit_product = profit_per_unit * total_quantity
            profit_margin = (profit_per_unit / avg_import_price * 100) if avg_import_price > 0 else 0
            
            profit_details.append({
                'product': product,
                'total_quantity': total_quantity,
                'avg_import_price': avg_import_price,
                'avg_export_price': avg_export_price,
                'profit_per_unit': profit_per_unit,
                'total_profit': total_profit_product,
                'profit_margin': profit_margin
            })
            
            total_export_value += total_export_value_product
            total_import_value += avg_import_price * total_quantity
    
    profit_details.sort(key=lambda x: x['total_profit'], reverse=True)
    total_profit = total_export_value - total_import_value
    profit_margin_total = (total_profit / total_import_value * 100) if total_import_value > 0 else 0
    
    # Tạo workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Báo cáo lợi nhuận"
    
    # Tiêu đề
    ws.merge_cells('A1:I1')
    ws['A1'] = "BÁO CÁO LỢI NHUẬN"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Thống kê tổng quan
    ws['A3'] = "TỔNG QUAN"
    ws['A3'].font = Font(bold=True, size=14)
    
    stats = [
        ['Chỉ tiêu', 'Giá trị'],
        ['Tổng giá trị nhập', total_import_value],
        ['Tổng giá trị xuất', total_export_value],
        ['Tổng lợi nhuận', total_profit],
        ['Tỷ lệ lợi nhuận', f"{profit_margin_total:.1f}%"],
    ]
    
    for row_idx, (label, value) in enumerate(stats, 4):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row_idx, column=2, value=value)
    
    # Header bảng chi tiết
    ws['A8'] = "CHI TIẾT LỢI NHUẬN THEO SẢN PHẨM"
    ws['A8'].font = Font(bold=True, size=14)
    
    headers = ['STT', 'Sản phẩm', 'Danh mục', 'Số lượng xuất', 'Giá nhập TB', 'Giá xuất TB', 'Lợi nhuận/SP', 'Tổng lợi nhuận', 'Tỷ lệ LN (%)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Dữ liệu chi tiết
    row = 11
    for item in profit_details:
        ws.cell(row=row, column=1, value=row-10)
        ws.cell(row=row, column=2, value=item['product'].name)
        ws.cell(row=row, column=3, value=item['product'].category.name)
        ws.cell(row=row, column=4, value=item['total_quantity'])
        ws.cell(row=row, column=5, value=item['avg_import_price'])
        ws.cell(row=row, column=6, value=item['avg_export_price'])
        ws.cell(row=row, column=7, value=item['profit_per_unit'])
        ws.cell(row=row, column=8, value=item['total_profit'])
        ws.cell(row=row, column=9, value=f"{item['profit_margin']:.1f}")
        row += 1
    
    # Tổng cộng
    ws.cell(row=row, column=1, value="TỔNG CỘNG").font = Font(bold=True)
    ws.cell(row=row, column=8, value=total_profit).font = Font(bold=True)
    ws.cell(row=row, column=9, value=f"{profit_margin_total:.1f}").font = Font(bold=True)
    
    # Điều chỉnh độ rộng cột
    adjust_column_width(ws)
    
    # Tạo response
    today = timezone.now().date()
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="bao_cao_loi_nhuan_{today.strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response
